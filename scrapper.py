
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from urllib.parse import urljoin, urlparse
import time
import random
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

class ContactScraper:
    def __init__(self):
        # Enhanced regex patterns
        self.EMAIL_REGEX = r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
        self.PHONE_REGEX = r"(\+?[1-9]\d{0,3}[\s.-]?\d{2,4}[\s.-]?\d{2,4}[\s.-]?\d{2,4})"
        self.WHATSAPP_REGEX = r"(?:whatsapp|whats app|wa|w\.a\.?)\s*:?\s*(\+?[1-9]\d{7,15})"
        self.URL_REGEX = r"https?://[^\s<>\"\'{}|\\^`\[\]]*"

        # Address keywords for better detection
        self.ADDRESS_KEYWORDS = [
            "address", "location", "addr", "office", "headquarters", "visit us", 
            "find us", "our location", "contact address", "postal address",
            "street", "building", "floor", "suite", "room", "block", "plot",
            "city", "state", "country", "pin", "zip", "postal code"
        ]

        # Setup session with retry strategy - Fixed compatibility issue
        self.session = requests.Session()

        # Try different parameter names for compatibility
        try:
            # New urllib3 version uses 'allowed_methods'
            retry_strategy = Retry(
                total=3,
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["HEAD", "GET", "OPTIONS"]
            )
        except TypeError:
            # Older urllib3 version uses 'method_whitelist'
            try:
                retry_strategy = Retry(
                    total=3,
                    status_forcelist=[429, 500, 502, 503, 504],
                    method_whitelist=["HEAD", "GET", "OPTIONS"]
                )
            except TypeError:
                # Fallback - basic retry without method restriction
                retry_strategy = Retry(
                    total=3,
                    status_forcelist=[429, 500, 502, 503, 504]
                )

        adapter = HTTPAdapter(max_retries=retry_strategy)
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)

        # Multiple user agents to rotate
        self.user_agents = [
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/119.0.0.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
        ]

    def get_headers(self):
        """Get randomized headers"""
        return {
            "User-Agent": random.choice(self.user_agents),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate, br",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Cache-Control": "max-age=0"
        }

    def clean_phone_numbers(self, raw_numbers):
        """Clean and filter phone numbers"""
        cleaned = set()
        for num in raw_numbers:
            # Remove extra spaces and normalize
            clean_num = re.sub(r'\s+', ' ', num.strip())
            digits_only = re.sub(r'\D', '', clean_num)

            # Skip if too short or too long
            if len(digits_only) < 7 or len(digits_only) > 15:
                continue

            # Skip obvious non-phone patterns
            if digits_only.startswith(('0000', '1111', '2222', '3333')):
                continue

            # Skip years, IDs, etc.
            if len(digits_only) == 4 and digits_only.startswith(('19', '20')):
                continue

            # Skip timestamps (10+ digits starting with 16/17/20)
            if len(digits_only) >= 10 and digits_only.startswith(('16', '17', '20')):
                continue

            cleaned.add(clean_num)
        return cleaned

    def extract_emails(self, soup):
        """Extract email addresses"""
        emails = set()

        # Extract from text content
        text = soup.get_text()
        found_emails = re.findall(self.EMAIL_REGEX, text, re.IGNORECASE)
        emails.update(found_emails)

        # Extract from mailto links
        for link in soup.find_all('a', href=True):
            href = link['href']
            if href.startswith('mailto:'):
                email = href.replace('mailto:', '').split('?')[0].split('&')[0]
                if re.match(self.EMAIL_REGEX, email):
                    emails.add(email)

        # Look in data attributes and other places
        for element in soup.find_all(attrs={'data-email': True}):
            email = element.get('data-email')
            if re.match(self.EMAIL_REGEX, email):
                emails.add(email)

        return emails

    def extract_phones(self, soup):
        """Extract phone numbers"""
        raw_phones = []

        # Extract from text content
        text = soup.get_text()
        text_phones = re.findall(self.PHONE_REGEX, text)
        raw_phones.extend(text_phones)

        # Extract from tel: links
        for link in soup.find_all('a', href=True):
            href = link['href']
            if href.startswith('tel:'):
                phone = href.replace('tel:', '').strip()
                raw_phones.append(phone)

        # Look for phone in data attributes
        for element in soup.find_all(attrs={'data-phone': True}):
            phone = element.get('data-phone')
            raw_phones.append(phone)

        return self.clean_phone_numbers(raw_phones)

    def extract_whatsapp(self, soup):
        """Extract WhatsApp numbers"""
        whatsapp_numbers = set()

        # Look for WhatsApp mentions with numbers in text
        text = soup.get_text()
        matches = re.findall(self.WHATSAPP_REGEX, text, re.IGNORECASE)
        for match in matches:
            clean_match = re.sub(r'\D', '', match)
            if len(clean_match) >= 7:
                whatsapp_numbers.add(match)

        # Check for WhatsApp links
        for link in soup.find_all('a', href=True):
            href = link['href']
            if 'wa.me' in href or 'whatsapp.com' in href or 'api.whatsapp.com' in href:
                # Extract number from WhatsApp URL
                wa_match = re.search(r'(\+?\d{7,15})', href)
                if wa_match:
                    whatsapp_numbers.add(wa_match.group(1))

        # Look for WhatsApp in button texts or data attributes
        whatsapp_keywords = ['whatsapp', 'whats app', 'wa', 'w.a.']
        for element in soup.find_all(['button', 'div', 'span'], string=re.compile('|'.join(whatsapp_keywords), re.I)):
            element_text = element.get_text()
            numbers = re.findall(r'(\+?\d[\d\s()-]{7,15})', element_text)
            for num in numbers:
                clean_num = re.sub(r'\D', '', num)
                if len(clean_num) >= 7:
                    whatsapp_numbers.add(num)

        return whatsapp_numbers

    def extract_urls(self, soup, base_url):
        """Extract website URLs"""
        urls = set()
        parsed_base = urlparse(base_url)
        base_domain = f"{parsed_base.scheme}://{parsed_base.netloc}"

        # Add the main website
        urls.add(base_domain)

        # Extract URLs from text content
        text = soup.get_text()
        text_urls = re.findall(self.URL_REGEX, text)
        for url in text_urls:
            # Clean URL - fixed syntax error
            punctuation_chars = '.,;:)]}>\"\''
            clean_url = url.rstrip(punctuation_chars)
            # Skip media files and resources
            if not clean_url.lower().endswith(('.jpg', '.png', '.css', '.js', '.pdf', '.jpeg', '.svg', '.gif', '.ico')):
                urls.add(clean_url)

        # Extract from href attributes
        for link in soup.find_all('a', href=True):
            href = link['href']
            if href.startswith('http'):
                if not href.lower().endswith(('.jpg', '.png', '.css', '.js', '.pdf', '.jpeg', '.svg', '.gif', '.ico')):
                    urls.add(href)
            elif href.startswith('/'):
                # Convert relative URLs to absolute
                full_url = urljoin(base_domain, href)
                urls.add(full_url)

        return urls

    def extract_addresses(self, soup):
        """Extract addresses using multiple strategies"""
        addresses = set()

        # Strategy 1: Look for <address> tags
        for addr_tag in soup.find_all('address'):
            addr_text = addr_tag.get_text(separator=' ', strip=True)
            if len(addr_text) > 10:
                addresses.add(addr_text)

        # Strategy 2: Look for elements with address-related class names
        address_classes = ['address', 'location', 'contact-info', 'office-address']
        for class_name in address_classes:
            for element in soup.find_all(attrs={'class': re.compile(class_name, re.I)}):
                addr_text = element.get_text(separator=' ', strip=True)
                if 15 < len(addr_text) < 400:
                    addresses.add(addr_text)

        # Strategy 3: Look for schema.org address markup
        for element in soup.find_all(attrs={'itemtype': re.compile('address', re.I)}):
            addr_text = element.get_text(separator=' ', strip=True)
            if len(addr_text) > 10:
                addresses.add(addr_text)

        # Strategy 4: Text-based detection with context
        text_lines = soup.get_text().split('\n')
        for i, line in enumerate(text_lines):
            line = line.strip()
            if line and any(keyword in line.lower() for keyword in self.ADDRESS_KEYWORDS):
                # Get surrounding context (previous and next lines)
                context_start = max(0, i - 1)
                context_end = min(len(text_lines), i + 4)
                context_lines = [l.strip() for l in text_lines[context_start:context_end] if l.strip()]
                context = ' '.join(context_lines)

                if 20 < len(context) < 500:
                    addresses.add(context)

        # Strategy 5: Look for postal codes and specific patterns
        postal_patterns = [
            r'\d{5,6}',  # 5-6 digit codes
            r'\d{3}\s*\d{3}',  # 3+3 digit codes
            r'[A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2}'  # UK style postcodes
        ]

        text = soup.get_text()
        for pattern in postal_patterns:
            matches = re.finditer(pattern, text)
            for match in matches:
                # Get context around postal code
                start_pos = max(0, match.start() - 100)
                end_pos = min(len(text), match.end() + 100)
                context = text[start_pos:end_pos].strip()

                # Check if it looks like an address
                if any(keyword in context.lower() for keyword in ['street', 'road', 'avenue', 'city', 'town']):
                    lines = context.split('\n')
                    address_lines = [line.strip() for line in lines if line.strip() and len(line.strip()) > 5]
                    if address_lines:
                        full_address = ' '.join(address_lines)
                        if 30 < len(full_address) < 300:
                            addresses.add(full_address)

        return addresses

    def scrape_contact_page(self, url):
        """Main scraping method"""
        # Normalize URL
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url

        print(f"Scraping: {url}")

        try:
            # Add random delay to be respectful
            time.sleep(random.uniform(0.5, 2.0))

            # Get page with custom headers
            headers = self.get_headers()
            response = self.session.get(url, headers=headers, timeout=20, allow_redirects=True)

            # Check if request was successful
            if response.status_code != 200:
                print(f"HTTP {response.status_code} for {url}")
                return self._empty_result()

            # Check content type
            content_type = response.headers.get('content-type', '')
            if 'text/html' not in content_type:
                print(f"Non-HTML content type: {content_type}")
                return self._empty_result()

            # Parse HTML
            soup = BeautifulSoup(response.content, 'html.parser')

            # Extract all data
            emails = self.extract_emails(soup)
            phones = self.extract_phones(soup)
            whatsapp = self.extract_whatsapp(soup)
            urls = self.extract_urls(soup, url)
            addresses = self.extract_addresses(soup)

            return {
                'emails': emails,
                'phones': phones,
                'whatsapp': whatsapp,
                'urls': urls,
                'addresses': addresses,
                'status': 'success'
            }

        except requests.exceptions.Timeout:
            print(f"Timeout error for {url}")
            return self._empty_result(status='timeout')
        except requests.exceptions.ConnectionError:
            print(f"Connection error for {url}")
            return self._empty_result(status='connection_error')
        except requests.exceptions.RequestException as e:
            print(f"Request error for {url}: {e}")
            return self._empty_result(status='request_error')
        except Exception as e:
            print(f"Unexpected error for {url}: {e}")
            return self._empty_result(status='unexpected_error')

    def _empty_result(self, status='failed'):
        """Return empty result structure"""
        return {
            'emails': set(),
            'phones': set(),
            'whatsapp': set(),
            'urls': set(),
            'addresses': set(),
            'status': status
        }

    def save_to_excel(self, url, data, filename="scraped_contacts.xlsx"):
        """Save results to Excel"""
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            # Create headers
            ws.append([
                "URL", 
                "Status",
                "Emails", 
                "Phone Numbers", 
                "WhatsApp Numbers", 
                "Website URLs", 
                "Addresses"
            ])

        # Prepare data for Excel
        row_data = [
            url,
            data.get('status', 'unknown'),
            " | ".join(data['emails']) if data['emails'] else "Not Found",
            " | ".join(data['phones']) if data['phones'] else "Not Found",
            " | ".join(data['whatsapp']) if data['whatsapp'] else "Not Found",
            " | ".join(list(data['urls'])[:3]) if data['urls'] else url,  # Limit URLs
            " | ".join(list(data['addresses'])[:2]) if data['addresses'] else "Not Found"  # Limit addresses
        ]

        ws.append(row_data)
        wb.save(filename)
        print(f"Results saved to: {filename}")

    def print_results(self, url, data):
        """Print results to console"""
        print(f"\n{'='*60}")
        print(f"RESULTS FOR: {url}")
        print(f"Status: {data.get('status', 'unknown')}")
        print(f"{'='*60}")

        print(f"📧 EMAILS ({len(data['emails'])}):")
        if data['emails']:
            for email in sorted(data['emails']):
                print(f"   • {email}")
        else:
            print("   • None found")

        print(f"\n📞 PHONE NUMBERS ({len(data['phones'])}):")
        if data['phones']:
            for phone in sorted(data['phones']):
                print(f"   • {phone}")
        else:
            print("   • None found")

        print(f"\n💬 WHATSAPP NUMBERS ({len(data['whatsapp'])}):")
        if data['whatsapp']:
            for wa in sorted(data['whatsapp']):
                print(f"   • {wa}")
        else:
            print("   • None found")

        print(f"\n🌐 WEBSITE URLS ({len(data['urls'])}):")
        if data['urls']:
            for url_item in sorted(list(data['urls'])[:5]):  # Show max 5 URLs
                print(f"   • {url_item}")
            if len(data['urls']) > 5:
                print(f"   • ... and {len(data['urls'])-5} more")
        else:
            print("   • None found")

        print(f"\n📍 ADDRESSES ({len(data['addresses'])}):")
        if data['addresses']:
            for i, addr in enumerate(sorted(list(data['addresses'])[:3]), 1):  # Show max 3 addresses
                print(f"   {i}. {addr[:100]}{'...' if len(addr) > 100 else ''}")
            if len(data['addresses']) > 3:
                print(f"   ... and {len(data['addresses'])-3} more")
        else:
            print("   • None found")

        print(f"{'='*60}")


# Main execution function
def main():
    scraper = ContactScraper()

    print("🔍 Contact Information Scraper")
    print("=" * 50)
    print("This tool extracts:")
    print("✓ Email addresses")
    print("✓ Phone numbers") 
    print("✓ WhatsApp numbers")
    print("✓ Website URLs")
    print("✓ Physical addresses")
    print("=" * 50)

    while True:
        url = input("\nEnter website URL (or 'exit' to quit): ").strip()

        if url.lower() == 'exit':
            print("\n👋 Thank you for using Contact Scraper!")
            break

        if not url:
            print("❌ Please enter a valid URL")
            continue

        try:
            # Scrape the website
            data = scraper.scrape_contact_page(url)

            # Display results
            scraper.print_results(url, data)

            # Save to Excel
            scraper.save_to_excel(url, data)

        except KeyboardInterrupt:
            print("\n⚠️ Operation cancelled by user")
            break
        except Exception as e:
            print(f"❌ Unexpected error: {e}")
            continue


if __name__ == "__main__":
    main()
